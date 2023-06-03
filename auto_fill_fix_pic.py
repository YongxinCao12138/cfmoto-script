import time

import openpyxl
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def main():
    excel_name = input("请输入文档名字")

    # 获取工作簿对象
    wb = openpyxl.load_workbook(excel_name)

    # 拿到第一个sheet
    sheet = wb.worksheets[0]

    # 获取工作表总行数
    rows = sheet.max_row
    tickets = []
    for row in range(rows):
        num = sheet.cell(row=row + 2, column=5).value
        if num is not None:
            ticket = {"row": row + 2, "num": num}
            tickets.append(ticket)
    print(tickets)

    # 创建一个Service对象
    service = Service(r'./chromedriver.exe')
    # service = Service(executable_path="./chromedriver")

    # 创建一个EdgeOptions对象，并设置一些选项
    options = webdriver.ChromeOptions()
    options.add_argument('window-size=1920x1080')  # 指定浏览器分辨率

    # 创建浏览器操作对象
    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()

    # 访问网页
    driver.get('https://dms.cfmoto-online.cn/cfdms/index.html')

    # 设置最大等待时间
    wait = WebDriverWait(driver, 5 * 60)

    print("please input username and password")

    # 售后
    wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div/div[2]/ul/li[2]/a"))).click()
    # 索赔单查询
    wait.until(EC.presence_of_element_located(
        (By.XPATH, "/html/body/div[1]/div[1]/div/div[2]/ul/li[2]/ul/li[2]/div/div/div/div/div[3]/ul/li[2]/a"))).click()
    print("done")

    for ticket in tickets:
        num = ticket["num"]
        row = ticket["row"]
        # 清除已经输入的索赔单号
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='in-oemClaimOrderMngQuery-mng-claimno']"))).clear()
        # 输入索赔单号
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='in-oemClaimOrderMngQuery-mng-claimno']"))).send_keys(
            num)
        # 搜索
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='btn-oemClaimOrderMngQuery-mng-search']"))).click()
        # 等待搜索按钮可点击后点击详情
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='btn-oemClaimOrderMngQuery-mng-search']")))
        try:
            element = driver.find_element(By.XPATH,
                                          "//*[@id='tab-claimOrderReportIndexQuery-list_content']/tbody/tr/td[3]/div/a")
            element.click()
        except NoSuchElementException:
            print("元素未找到")
        # 点击维修照片预览
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='dia-tab-work-order-previewPic']"))).click()
        # 复制图片
        # 定位图片元素
        image_element = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='dowebok']/li[1]/img")))

        # 点击图片展开
        image_element.click()

        time.sleep(30)
        # 获得展开图片元素,使用 Selenium 截图功能保存图片
        filename = f'{row}_{num}.png'
        wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='dia-filesView']/div/div[1]/img"))).screenshot(
            filename)

        print('图片已保存：', filename)

        # 关闭图片
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='dia-filesView']/div/div[4]"))).click()

        # 点击返回
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='btn-claimOrderDetail-changeToClaim-back']"))).click()

    # 关闭浏览器
    driver.quit()

    print("Done")


# 增加调用main()函数
if __name__ == '__main__':
    main()
